import openpyxl


def get_row_count(excel_file, sheet_name):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row


def get_column_count(excel_file, sheet_name):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_column


def read_data(excel_file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_num, column=col_num).value


def write_data(excel_file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(excel_file)
